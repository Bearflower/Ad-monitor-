from __future__ import annotations

import json
from collections import defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


DOWNLOAD_DIR = Path(
    "/Users/yl/Library/Application Support/ziniaobrowserdatas/"
    "ziniao browser/虾皮泰国"
)
OUTPUT_PATH = Path("var/imports/shopee/merged-order-cost-data.json")

KNOWN_COSTS = {
    "Foot Soak Bag-one bag": 5,
    "Foot Soak Bag-two bags": 11,
    "Foot Soak Bag-three bags": 17,
}


def normalize_status(value: object) -> str:
    text = str(value or "").strip()
    if text == "ยกเลิกแล้ว":
        return "已取消"
    if text == "สำเร็จแล้ว":
        return "已完成"
    if text == "จัดส่งสำเร็จแล้ว":
        return "配送完成"
    if text == "การจัดส่ง":
        return "运输中"
    if text == "ที่ต้องจัดส่ง":
        return "待发货"
    if text.startswith("ผู้ซื้อได้รับสินค้าแล้ว"):
        return "买家已收货（售后期内）"
    return text or "未知"


def load_rows(path: Path) -> list[dict[str, object]]:
    # Shopee exports an incorrect A1 dimension, so normal (non-streaming) mode
    # is required to recover all populated rows.
    sheet = load_workbook(path, read_only=False, data_only=True).active
    headers = [sheet.cell(1, col).value for col in range(1, sheet.max_column + 1)]
    return [
        dict(
            zip(
                headers,
                [
                    sheet.cell(row, col).value
                    for col in range(1, sheet.max_column + 1)
                ],
                strict=True,
            )
        )
        for row in range(2, sheet.max_row + 1)
    ]


def main() -> None:
    files = sorted(DOWNLOAD_DIR.glob("Order.all.*.xlsx"))
    if len(files) != 4:
        raise SystemExit(f"Expected 4 Shopee exports, found {len(files)}")

    source_rows: list[tuple[str, dict[str, object], list[str]]] = []
    file_summary = []
    for path in files:
        rows = load_rows(path)
        headers = list(rows[0]) if rows else []
        source_rows.extend((path.name, row, headers) for row in rows)
        file_summary.append(
            {
                "file": path.name,
                "rows": len(rows),
                "orders": len({str(row[headers[0]]) for row in rows}),
                "min_date": min(str(row[headers[6]])[:10] for row in rows),
                "max_date": max(str(row[headers[6]])[:10] for row in rows),
            }
        )

    seen: set[tuple[str, str, str, str]] = set()
    order_lines = []
    sku_rollup: dict[str, dict[str, object]] = defaultdict(
        lambda: {
            "names": set(),
            "variations": set(),
            "latest_name": "",
            "latest_name_date": "",
            "orders": set(),
            "quantity": 0,
            "returned_quantity": 0,
            "first_date": "9999-12-31",
            "last_date": "",
        }
    )

    for source_file, row, headers in source_rows:
        order_id = str(row[headers[0]] or "").strip()
        seller_sku = str(row[headers[19]] or "").strip()
        variation = str(row[headers[20]] or "").strip()
        product_name = str(row[headers[18]] or "").strip()
        key = (order_id, seller_sku, variation, product_name)
        if key in seen:
            continue
        seen.add(key)

        raw_status = str(row[headers[1]] or "").strip()
        order_date = str(row[headers[6]] or "")[:10]
        quantity = int(float(row[headers[23]] or 0))
        returned_quantity = int(float(row[headers[24]] or 0))
        tracking_number = str(row[headers[14]] or "").strip()
        shipping_method = str(row[headers[13]] or "").strip()
        normalized_status = normalize_status(raw_status)

        order_lines.append(
            {
                "order_date": order_date,
                "order_id": order_id,
                "order_status": normalized_status,
                "raw_order_status": raw_status,
                "shipping_status": (
                    "已取消"
                    if normalized_status == "已取消"
                    else "已有物流单号"
                    if tracking_number
                    else normalized_status
                ),
                "shipping_method": shipping_method,
                "tracking_number": tracking_number,
                "seller_sku": seller_sku,
                "product_name": product_name,
                "variation": variation,
                "quantity": quantity,
                "returned_quantity": returned_quantity,
                "source_file": source_file,
            }
        )

        if normalized_status == "已取消":
            continue
        rollup = sku_rollup[seller_sku]
        rollup["names"].add(product_name)
        rollup["variations"].add(variation)
        if order_date >= str(rollup["latest_name_date"]):
            rollup["latest_name"] = product_name
            rollup["latest_name_date"] = order_date
        rollup["orders"].add(order_id)
        rollup["quantity"] += quantity
        rollup["returned_quantity"] += returned_quantity
        rollup["first_date"] = min(str(rollup["first_date"]), order_date)
        rollup["last_date"] = max(str(rollup["last_date"]), order_date)

    sku_rows = []
    for seller_sku, values in sku_rollup.items():
        known_cost = KNOWN_COSTS.get(seller_sku)
        variations = sorted(values["variations"])
        sku_rows.append(
            {
                "platform": "Shopee",
                "store": "no4kud44da",
                "seller_sku": seller_sku,
                "variation": " / ".join(variations),
                "product_name": values["latest_name"],
                "first_date": values["first_date"],
                "last_date": values["last_date"],
                "order_count": len(values["orders"]),
                "quantity": values["quantity"],
                "returned_quantity": values["returned_quantity"],
                "net_quantity": values["quantity"] - values["returned_quantity"],
                "cost_status": "已填写" if known_cost is not None else "待填写",
                "unit_cost_cny": known_cost,
                "cost_effective_date": values["first_date"],
                "cost_note": "历史已确认成本" if known_cost is not None else None,
            }
        )

    sku_rows.sort(
        key=lambda item: (
            item["unit_cost_cny"] is not None,
            -int(item["quantity"]),
            str(item["seller_sku"]),
        )
    )
    order_lines.sort(key=lambda item: (item["order_date"], item["order_id"]))

    payload = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "store": "no4kud44da",
        "file_summary": file_summary,
        "source_row_count": len(source_rows),
        "deduplicated_line_count": len(order_lines),
        "unique_order_count": len({line["order_id"] for line in order_lines}),
        "noncancel_order_count": len(
            {
                line["order_id"]
                for line in order_lines
                if line["order_status"] != "已取消"
            }
        ),
        "sku_count": len(sku_rows),
        "pending_cost_count": sum(
            1 for item in sku_rows if item["unit_cost_cny"] is None
        ),
        "sku_rows": sku_rows,
        "order_lines": order_lines,
    }
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT_PATH.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    print(json.dumps({key: payload[key] for key in (
        "source_row_count",
        "deduplicated_line_count",
        "unique_order_count",
        "noncancel_order_count",
        "sku_count",
        "pending_cost_count",
    )}, ensure_ascii=False))


if __name__ == "__main__":
    main()
