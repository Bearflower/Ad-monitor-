from __future__ import annotations

from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

from adwatch.analytics.business_inputs import BusinessInputError
from adwatch.orders.repository import OrderRepository
from adwatch.storage.db import Database

HEADERS = (
    "平台", "店铺", "商品名称", "Item ID", "Model ID", "Seller SKU",
    "规格", "当前库存", "首次发现日期", "最近销售日期", "待匹配订单数",
    "待匹配件数", "成本状态", "数据来源", "单件成本_人民币",
    "成本生效日期", "成本备注",
)


def export_pending_sku_costs(database: Database, destination: Path) -> int:
    rows = OrderRepository(database).pending_sku_costs()
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "待补成本"
    sheet.append(HEADERS)
    for row in rows:
        sheet.append(
            (
                row["platform"], row["store"], row["product_name"],
                row["item_id"], row["model_id"], row["seller_sku"],
                row["variation_name"], row["inventory_units"],
                row["first_seen_date"], row["latest_order_date"] or "",
                row["pending_orders"], row["pending_units"], "待填写",
                "紫鸟CLI", "", row["first_seen_date"], "",
            )
        )
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:Q{max(sheet.max_row, 1)}"
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4472C4")
    for row in range(2, sheet.max_row + 1):
        for column in range(15, 18):
            sheet.cell(row, column).fill = PatternFill(
                "solid", fgColor="FFF2CC"
            )
    widths = (10, 14, 36, 16, 16, 28, 16, 12, 14, 14, 14, 14, 12, 12, 18, 16, 24)
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[chr(64 + index)].width = width
    destination.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(destination)
    return len(rows)


def import_sku_costs(database: Database, source: Path) -> int:
    workbook = load_workbook(source, data_only=True)
    if "待补SKU成本" in workbook.sheetnames:
        sheet = workbook["待补SKU成本"]
        header_row = 5
        expected = (
            "平台", "店铺", "Seller SKU", "规格", "商品名称",
            "首次订单日期", "最近订单日期", "相关订单数", "销售件数",
            "退货件数", "净件数", "成本状态", "单件成本（人民币）",
            "成本生效日期", "成本备注",
        )
        indexes = (0, 1, 2, 12, 13, 14)
    else:
        sheet = workbook["待补成本"]
        header_row = 1
        expected = HEADERS
        indexes = (0, 1, 5, 14, 15, 16)
    headers = [cell.value for cell in sheet[header_row]]
    if tuple(headers[: len(expected)]) != expected:
        raise BusinessInputError("待补成本表头不匹配")
    values = []
    for line, row in enumerate(
        sheet.iter_rows(min_row=header_row + 1, values_only=True),
        header_row + 1,
    ):
        platform_index, store_index, sku_index, cost_index, date_index, note_index = indexes
        raw_cost = row[cost_index]
        if raw_cost in (None, ""):
            continue
        try:
            cost = Decimal(str(raw_cost))
            effective = _as_date(row[date_index])
        except (InvalidOperation, ValueError) as error:
            raise BusinessInputError(f"第 {line} 行成本或日期无效") from error
        required = (platform_index, store_index, sku_index)
        if cost < 0 or not all(str(row[i] or "").strip() for i in required):
            raise BusinessInputError(f"第 {line} 行必填值无效")
        values.append(
            (
                str(row[platform_index]).strip().lower(),
                str(row[store_index]).strip(),
                str(row[sku_index]).strip(),
                effective.isoformat(),
                str(cost),
                str(row[note_index] or "").strip(),
            )
        )
    with database.transaction() as connection:
        for value in values:
            connection.execute(
                """
                INSERT INTO sku_cost_history(
                    platform, store, seller_sku, effective_date,
                    unit_cost_cny, note
                ) VALUES (?, ?, ?, ?, ?, ?)
                ON CONFLICT(platform, store, seller_sku, effective_date)
                DO UPDATE SET
                    unit_cost_cny=excluded.unit_cost_cny,
                    note=excluded.note,
                    updated_at=strftime('%Y-%m-%dT%H:%M:%fZ', 'now')
                """,
                value,
            )
    return len(values)


def _as_date(value: object) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return date.fromisoformat(str(value).strip())
