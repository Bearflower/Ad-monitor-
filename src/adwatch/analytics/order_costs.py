from __future__ import annotations

import csv
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook

from adwatch.analytics.business_inputs import BusinessInputError
from adwatch.storage.db import Database

HEADERS = ("日期", "平台", "店铺", "订单号", "SKU", "数量", "单件成本_人民币")
PLATFORMS = {"shopee", "tiktok"}
CENT = Decimal("0.01")


@dataclass(frozen=True)
class OrderCostLine:
    platform: str
    store: str
    order_id: str
    sku_id: str
    order_date: date
    quantity: int
    unit_cost_cny: Decimal

    @property
    def line_cost_cny(self) -> Decimal:
        return self.unit_cost_cny * self.quantity

    @property
    def key(self) -> tuple[str, str, str, str]:
        return self.platform, self.store, self.order_id, self.sku_id


@dataclass(frozen=True)
class OrderImportSummary:
    read: int
    inserted: int
    updated: int
    deduplicated: int
    start: date
    end: date
    total_cost_cny: Decimal


def import_order_costs(database: Database, source: Path) -> OrderImportSummary:
    raw = _raw_rows(source)
    if not raw:
        raise BusinessInputError("no order rows")
    parsed = [_parse_line(row, line) for line, row in raw]
    unique: dict[tuple[str, str, str, str], OrderCostLine] = {}
    deduplicated = 0
    for item in parsed:
        existing = unique.get(item.key)
        if existing is None:
            unique[item.key] = item
        elif existing == item:
            deduplicated += 1
        else:
            raise BusinessInputError(
                f"conflicting duplicate: {'/'.join(item.key)}"
            )

    inserted = 0
    updated = 0
    with database.transaction() as connection:
        for item in unique.values():
            exists = connection.execute(
                """
                SELECT 1 FROM order_cost_lines
                WHERE platform=? AND store=? AND order_id=? AND sku_id=?
                """,
                item.key,
            ).fetchone()
            connection.execute(
                """
                INSERT INTO order_cost_lines(
                    platform, store, order_id, sku_id, order_date,
                    quantity, unit_cost_cny, line_cost_cny, source_file
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                ON CONFLICT(platform, store, order_id, sku_id) DO UPDATE SET
                    order_date=excluded.order_date,
                    quantity=excluded.quantity,
                    unit_cost_cny=excluded.unit_cost_cny,
                    line_cost_cny=excluded.line_cost_cny,
                    source_file=excluded.source_file,
                    updated_at=strftime('%Y-%m-%dT%H:%M:%fZ', 'now')
                """,
                (
                    item.platform,
                    item.store,
                    item.order_id,
                    item.sku_id,
                    item.order_date.isoformat(),
                    item.quantity,
                    str(item.unit_cost_cny),
                    str(item.line_cost_cny),
                    source.name,
                ),
            )
            inserted += int(exists is None)
            updated += int(exists is not None)

    dates = [item.order_date for item in unique.values()]
    total = sum(
        (item.line_cost_cny for item in unique.values()),
        Decimal("0"),
    )
    return OrderImportSummary(
        read=len(raw),
        inserted=inserted,
        updated=updated,
        deduplicated=deduplicated,
        start=min(dates),
        end=max(dates),
        total_cost_cny=total.quantize(CENT),
    )


def _raw_rows(path: Path) -> list[tuple[int, dict[str, object]]]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        with path.open(encoding="utf-8-sig", newline="") as handle:
            reader = csv.DictReader(handle)
            _require_headers(reader.fieldnames)
            return [
                (line, dict(row))
                for line, row in enumerate(reader, 2)
                if any(value not in (None, "") for value in row.values())
            ]
    if suffix == ".xlsx":
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            for sheet in workbook.worksheets:
                values = sheet.iter_rows(values_only=True)
                header = next(values, None)
                if header is None:
                    continue
                names = tuple(
                    "" if value is None else str(value).strip()
                    for value in header
                )
                _require_headers(names)
                return [
                    (line, dict(zip(names, row, strict=False)))
                    for line, row in enumerate(values, 2)
                    if any(value not in (None, "") for value in row)
                ]
        finally:
            workbook.close()
        raise BusinessInputError("workbook has no non-empty worksheet")
    raise BusinessInputError("file must be .csv or .xlsx")


def _require_headers(fieldnames: Iterable[str] | None) -> None:
    names = tuple(fieldnames or ())
    missing = [name for name in HEADERS if name not in names]
    if missing:
        raise BusinessInputError(f"missing columns: {', '.join(missing)}")


def _parse_line(row: dict[str, object], line: int) -> OrderCostLine:
    try:
        order_date = _parse_date(row.get("日期"))
        platform = _text(row.get("平台")).lower()
        store = _text(row.get("店铺"))
        order_id = _text(row.get("订单号"))
        sku_id = _text(row.get("SKU"))
        quantity = _positive_integer(row.get("数量"))
        unit_cost = _cost(row.get("单件成本_人民币"))
    except (InvalidOperation, TypeError, ValueError) as error:
        raise BusinessInputError(f"line {line} has invalid values") from error
    if platform not in PLATFORMS:
        raise BusinessInputError(f"line {line} has unsupported platform")
    return OrderCostLine(
        platform=platform,
        store=store,
        order_id=order_id,
        sku_id=sku_id,
        order_date=order_date,
        quantity=quantity,
        unit_cost_cny=unit_cost,
    )


def _parse_date(value: object) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, bool):
        raise ValueError("boolean date")
    if isinstance(value, (int, float, Decimal)):
        decimal = Decimal(str(value))
        if decimal != decimal.to_integral_value():
            raise ValueError("fractional date")
        text = str(int(decimal))
    else:
        text = _text(value)
    if len(text) == 8 and text.isdigit():
        return datetime.strptime(text, "%Y%m%d").date()
    return date.fromisoformat(text)


def _text(value: object) -> str:
    if value is None or isinstance(value, bool):
        raise ValueError("missing text")
    text = str(value).strip()
    if not text:
        raise ValueError("missing text")
    return text


def _positive_integer(value: object) -> int:
    if value is None or isinstance(value, bool):
        raise ValueError("invalid quantity")
    number = Decimal(str(value).strip())
    if number != number.to_integral_value() or number <= 0:
        raise ValueError("invalid quantity")
    return int(number)


def _cost(value: object) -> Decimal:
    if value is None or isinstance(value, bool):
        raise ValueError("invalid cost")
    number = Decimal(str(value).strip())
    if not number.is_finite() or number < 0 or number.as_tuple().exponent < -4:
        raise ValueError("invalid cost")
    return number
