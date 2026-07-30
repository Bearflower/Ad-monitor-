from __future__ import annotations

import hashlib
from collections import defaultdict
from datetime import datetime, time
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook

from adwatch.orders.models import ParseResult, PlatformOrderLine, PlatformSku

_HEADERS = {
    "order_id": "หมายเลขคำสั่งซื้อ",
    "order_status": "สถานะการสั่งซื้อ",
    "refund_status": "สถานะการคืนเงินหรือคืนสินค้า",
    "ordered_at": "วันที่ทำการสั่งซื้อ",
    "product_name": "ชื่อสินค้า",
    "seller_sku": "เลขอ้างอิง SKU (SKU Reference No.)",
    "variation_name": "ชื่อตัวเลือก",
    "quantity": "จำนวน",
    "buyer_paid": "ราคาสินค้าที่ชำระโดยผู้ซื้อ (THB)",
}

_STATUS = {
    "ยกเลิกแล้ว": "cancelled",
    "สำเร็จแล้ว": "completed",
    "จัดส่งสำเร็จแล้ว": "delivered",
    "การจัดส่ง": "shipped",
    "ที่ต้องจัดส่ง": "pending",
}


def _stable_id(prefix: str, value: str) -> str:
    digest = hashlib.sha256(value.encode("utf-8")).hexdigest()[:16]
    return f"{prefix}-{digest}"


def _decimal(value: object) -> Decimal:
    if value in (None, ""):
        return Decimal("0")
    return Decimal(str(value).replace(",", ""))


def parse_order_export(
    source: Path,
    *,
    store: str,
    source_updated_at: datetime,
) -> ParseResult:
    sheet = load_workbook(source, read_only=False, data_only=True).active
    headers = [
        str(sheet.cell(1, column).value or "").strip()
        for column in range(1, sheet.max_column + 1)
    ]
    positions = {header: index for index, header in enumerate(headers)}
    missing = [
        header for header in _HEADERS.values() if header not in positions
    ]
    if missing:
        raise ValueError(
            "Shopee order export missing columns: " + ", ".join(missing)
        )

    grouped: dict[tuple[str, str], dict[str, object]] = defaultdict(
        lambda: {"quantity": 0, "buyer_paid": Decimal("0")}
    )
    rejected: list[str] = []
    for row_number, values in enumerate(
        sheet.iter_rows(min_row=2, values_only=True), start=2
    ):
        row = {
            key: values[positions[header]]
            for key, header in _HEADERS.items()
        }
        order_id = str(row["order_id"] or "").strip()
        seller_sku = str(row["seller_sku"] or "").strip()
        if not order_id or not seller_sku:
            rejected.append(
                f"row {row_number}: missing order ID or Seller SKU"
            )
            continue
        key = (order_id, seller_sku)
        item = grouped[key]
        previous_quantity = int(item["quantity"])
        previous_buyer_paid = Decimal(str(item["buyer_paid"]))
        item.update(row)
        item["quantity"] = previous_quantity + int(_decimal(row["quantity"]))
        item["buyer_paid"] = previous_buyer_paid + _decimal(row["buyer_paid"])

    orders: list[PlatformOrderLine] = []
    sku_by_key: dict[tuple[str, str], PlatformSku] = {}
    for (order_id, seller_sku), row in grouped.items():
        product_name = str(row["product_name"] or "").strip()
        variation_name = str(row["variation_name"] or "").strip()
        item_id = _stable_id("product", product_name)
        model_id = _stable_id("sku", seller_sku)
        raw_status = str(row["order_status"] or "").strip()
        order_status = _STATUS.get(raw_status, raw_status.lower() or "pending")
        raw_refund = str(row["refund_status"] or "").strip()
        refund_status = (
            "returned" if raw_refund else ""
        )
        ordered_at = datetime.fromisoformat(
            str(row["ordered_at"]).strip()
        ).date()
        logistics_status = (
            order_status
            if order_status in {"shipped", "delivered", "completed"}
            else "pending"
        )
        orders.append(
            PlatformOrderLine(
                platform="shopee",
                store=store,
                order_id=order_id,
                item_id=item_id,
                model_id=model_id,
                seller_sku=seller_sku,
                variation_name=variation_name,
                product_name=product_name,
                quantity=int(row["quantity"]),
                buyer_paid=Decimal(str(row["buyer_paid"])),
                currency="THB",
                order_status=order_status,
                logistics_status=logistics_status,
                refund_status=refund_status,
                ordered_at=ordered_at,
                source_updated_at=source_updated_at,
            )
        )
        observed_at = datetime.combine(
            ordered_at, time.min, tzinfo=source_updated_at.tzinfo
        )
        existing_sku = sku_by_key.get((item_id, model_id))
        sku = PlatformSku(
            platform="shopee",
            store=store,
            item_id=item_id,
            model_id=model_id,
            seller_sku=seller_sku,
            variation_name=variation_name,
            product_name=product_name,
            inventory_units=0,
            observed_at=observed_at,
        )
        if existing_sku is None or observed_at < existing_sku.observed_at:
            sku_by_key[(item_id, model_id)] = sku
    return ParseResult(
        orders=tuple(orders),
        skus=tuple(sku_by_key.values()),
        rejected=tuple(rejected),
    )
