from datetime import datetime
from decimal import Decimal

from openpyxl import load_workbook
from openpyxl import Workbook

from adwatch.analytics.sku_cost_workbook import (
    export_pending_sku_costs,
    import_sku_costs,
)
from adwatch.orders.models import PlatformSku
from adwatch.orders.repository import OrderRepository
from adwatch.storage.db import Database


def test_pending_workbook_prefills_facts_and_imports_only_cost(tmp_path):
    database = Database(tmp_path / "test.sqlite3")
    database.migrate()
    OrderRepository(database).upsert_skus(
        (
            PlatformSku(
                "shopee", "虾皮泰国", "57861884313", "311033956020",
                "Foot Soak Bag-one bag", "1 bag", "泡脚包", 31,
                datetime(2026, 7, 28, 9),
            ),
        )
    )
    target = tmp_path / "pending.xlsx"

    assert export_pending_sku_costs(database, target) == 1
    workbook = load_workbook(target)
    sheet = workbook["待补成本"]
    assert sheet.freeze_panes == "A2"
    assert sheet["F2"].value == "Foot Soak Bag-one bag"
    assert sheet["H2"].value == 31

    sheet["O2"] = 5
    sheet["P2"] = "2026-07-08"
    workbook.save(target)
    assert import_sku_costs(database, target) == 1

    with database.connect() as connection:
        row = connection.execute("SELECT * FROM sku_cost_history").fetchone()
    assert Decimal(row["unit_cost_cny"]) == Decimal("5")
    assert row["effective_date"] == "2026-07-08"


def test_imports_simplified_user_facing_cost_workbook(tmp_path):
    database = Database(tmp_path / "test.sqlite3")
    database.migrate()
    target = tmp_path / "pending-simple.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "待补SKU成本"
    sheet.append(["说明"])
    sheet.append([])
    sheet.append([])
    sheet.append([])
    sheet.append([
        "平台", "店铺", "Seller SKU", "规格", "商品名称",
        "首次订单日期", "最近订单日期", "相关订单数", "销售件数",
        "退货件数", "净件数", "成本状态", "单件成本（人民币）",
        "成本生效日期", "成本备注",
    ])
    sheet.append([
        "Shopee", "no4kud44da", "0500106001010", "Black,140cm",
        "Leather Belt", "2026-04-10", "2026-05-07", 14, 14, 1, 13,
        "已填写", 12.5, "2026-04-10", "首批订单成本",
    ])
    workbook.save(target)

    assert import_sku_costs(database, target) == 1

    with database.connect() as connection:
        row = connection.execute("SELECT * FROM sku_cost_history").fetchone()
    assert row["platform"] == "shopee"
    assert row["store"] == "no4kud44da"
    assert row["seller_sku"] == "0500106001010"
    assert Decimal(row["unit_cost_cny"]) == Decimal("12.5")
    assert row["effective_date"] == "2026-04-10"
